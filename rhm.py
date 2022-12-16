from re import template
from flask import Flask, render_template, request, session
import json
import random
import os
import openpyxl
from player_class import Player
from game_session_class import GameSession
from mini_game_class import Minigame

rhm_site = Flask(__name__)
rhm_site.config["SESSION_PERMANENT"] = False
rhm_site.config["SESSION_TYPE"] = "filesystem"

global_player_data = []
global_file_data = []
global_game_category = dict() # key (category) with value list [amount, color]
global_probability_data = []
global_settings_data = [(0, 0, 0, 1)]
global_game_session = None
global_current_game = None
global_skipped_game = None

clean_player_list = list()

@rhm_site.route("/")
@rhm_site.route("/start_page")
def game_config_page():
    global global_player_data
    global_player_data = []
    global global_file_data
    global_file_data = []
    global global_game_category
    global_game_category = dict()
    global global_probability_data
    global_probability_data = []
    global global_skipped_game
    global_skipped_game = None

    global clean_player_list
    clean_player_list = list()
    return render_template('start_page.html')

@rhm_site.route("/initialize_players", methods=['GET'])
def initialize_players():
    return render_template('initialize_players.html')

@rhm_site.route("/import_spreadsheet", methods=['GET'])
def import_spreadsheet():
    BASE_DIR = 'C:\\Users\\siniz\\AppData\\Local\\Programs\\Python\\Python38\\rhythm multiplayer 2\\static\\rhythm_games'
    files = os.listdir(BASE_DIR)
    filtered_files = [file for file in files if file.endswith('.xlsx')]
    global global_file_data
    filtered_files_with_count = [(file, get_game_count(file), "") for file in filtered_files]

    # data sent is a list of tuples. each tuple is (file_name, game_amt_in_each_file, ""/ " disabled")
    if len(global_file_data) > 0:
        filtered_files_with_count = [(file, get_game_count(file), "" if file in global_file_data else " disabled") for file in filtered_files]
    
    return render_template('import_spreadsheet.html', files = filtered_files_with_count)

@rhm_site.route("/calculate_categories", methods=['GET'])
def calculate_category():    
    #convert global_file_data to list
    global global_file_data
    global global_game_category
    if (len(global_game_category.keys()) == 0):
        global_game_category = dict()
        file_list = convert_to_list(global_file_data)   
        for file in file_list:
            add_game_categories(file)
    
    
    category_list = list(global_game_category.keys())
    sent_category_list = list()
    probability_amt = 100 // len(category_list)
    global global_probability_data
    user_prob_exists = len(global_probability_data) > 0
    # information getting sent is a list of tuples - each tuple is (category_name, number_of_games_in_each_category, color_for_category,
    # probability, is_enabled)

    for i in range(len(category_list)):
        sent_prob_amt = probability_amt
        category_name = category_list[i]
        if (i == 0 and probability_amt != 100.0 / len(category_list)):
            sent_prob_amt+= (100 % len(category_list))
        if user_prob_exists:
            prob_data = convert_to_list(global_probability_data)
            sent_prob_amt = int(prob_data[i].split(':')[1])
        sent_category_list.append((category_name, global_game_category[category_name][0], global_game_category[category_name][1], sent_prob_amt, sent_prob_amt > -1))
    return render_template('calculate_categories.html', values = sent_category_list)

@rhm_site.route("/game_settings")
def game_settings():
    global global_settings_data
    return render_template('game_settings.html', values = global_settings_data)

@rhm_site.route("/show_player_info")
def show_player_info():
    global global_player_data
    temp_list = convert_player_json_to_list_tuple(global_player_data)
    store_player_list_nicer(temp_list)
    data_tuple = convert_list_to_display(temp_list)
    initialize_game_session()
    return render_template('show_player_info.html', dataLeft = data_tuple[0], dataRight = data_tuple[1])
    # return render_template('show_player_info.html')

@rhm_site.route("/round_turn", methods=['GET', 'POST'])
def round_turn():
    global global_skipped_game
    if request.method == 'POST':
        game_chosen = request.form['skipped_game_identifiers'] # string needs to be converted to minigame type (done in game session)
        point_value = request.form['point_value']
        if point_value != None and point_value != '':
            global global_game_session
            global global_current_game
            global_game_session.determine_point_value_and_add_to_data(global_current_game, int(point_value))
            global_skipped_game = None
            data = {'point_value': point_value}
            return jsonify(data)
        print(game_chosen)
        global_skipped_game = game_chosen
        return game_chosen
    game_list, point_list, chosen_players, next_player = get_game_turn_data(global_skipped_game)
    if len(chosen_players[0]) == 5:
        return render_template('round_turn_2p.html') # need to fix 2p html
    return render_template('round_turn.html', currPlayer = chosen_players, game = game_list, pointData = point_list, nextPlayer = next_player, round=global_game_session.get_round(), turn = global_game_session.get_remaining_players_in_round()) # using non point system rn

# delete this later - for testing only
@rhm_site.route("/round_turn_2p", methods=['GET', 'POST'])
def round_turn_2p():
    return render_template('round_turn_2p.html')

# this route is for testing only!
@rhm_site.route("/show_info", methods=['GET'])
def show_info():
    global global_settings_data
    print("global")
    print(global_settings_data)
    return render_template('show_info.html', data=global_settings_data)

# POST methods

@rhm_site.route("/player_data", methods=['POST'])
def player_data():
    player_data = request.form['body']
    global global_player_data
    global_player_data = player_data
    return player_data

@rhm_site.route("/file_data", methods=['POST'])
def file_data():
    file_data = request.form['body']
    global global_file_data
    if global_file_data != file_data:
        global global_game_category
        global_game_category = dict()
        global global_probability_data
        global_probability_data = []
        global_file_data = file_data
    return file_data

@rhm_site.route("/probability_data", methods=['POST'])
def probability_data():
    probability_data = request.form['body']
    total_games = request.form['total']
    global global_probability_data
    if (global_probability_data != probability_data):
        global_probability_data = probability_data
        global global_settings_data
        global_settings_data = [(0, 0, total_games, 1)]
    
    return probability_data

@rhm_site.route("/settings_data", methods=['POST'])
def settings_data():
    settings_data = request.form['body']
    global global_settings_data

    #convert into tuple in list
    temp_list = list()
    info_list = convert_to_list(settings_data)
    temp_list.append(tuple(info_list))
    print(temp_list)
    global_settings_data = temp_list

    return settings_data
    


# HELPER functions
def get_game_count(file_name: str):
    BASE_DIR = 'C:\\Users\\siniz\\AppData\\Local\\Programs\\Python\\Python38\\rhythm multiplayer 2\\static\\rhythm_games'
    full_file_name = BASE_DIR + '\\' + file_name
    game_count = 0
    wookbook = openpyxl.load_workbook(full_file_name)

    # Define variable to read the active sheet:
    worksheet = wookbook.active

    # Iterate the loop to read the cell values
    for i in range(1, worksheet.max_row):
        for col in worksheet.iter_cols(1, 1):
            if col[i].value != None:
                game_count += 1
    return game_count

def add_game_categories(file_name:str):
    BASE_DIR = 'C:\\Users\\siniz\\AppData\\Local\\Programs\\Python\\Python38\\rhythm multiplayer 2\\static\\rhythm_games'
    full_file_name = BASE_DIR + '\\' + file_name
    global global_game_category

    wookbook = openpyxl.load_workbook(full_file_name)

    # Define variable to read the active sheet:
    worksheet = wookbook.active

    # Iterate the loop to read the cell values
    for i in range(1, worksheet.max_row):
        for col in worksheet.iter_cols(2, 2):
            if col[i].value != None:
                category_val = col[i].value
                if category_val in global_game_category:
                    global_game_category[category_val][0]+=1
                else:
                    global_game_category[category_val] = list()
                    global_game_category[category_val].append(1)
                    global_game_category[category_val].append(col[i].fill.bgColor.index[2:])

def convert_to_list(strList: str):
    split_list = strList.replace('"', '').replace('[', '').replace(']', '').split(',')
    return split_list

def convert_player_json_to_list_tuple(strList: str):
    split_list = strList.replace('"', '').replace('[', '').replace(']', '').split(',')
    final_list = list()
    new_list = list()
    for item in split_list:
        if item.startswith('{'):
            if len(new_list) != 0:
                
                final_list.append(tuple(new_list))
            new_list = list()
        value = item.split(":")[1].replace("}", "")
        new_list.append(value)
    if len(new_list) == 3:
        final_list.append(tuple(new_list))
    
    random.shuffle(final_list)

    return final_list

def convert_list_to_display(final_list):
    left_list = list()
    right_list = list()
    for x in range(min(len(final_list), 6)):
        temp = list(final_list[x])
        temp.insert(0, " " + str(x+1))
        left_list.append(tuple(temp))
    for x in range(len(left_list), 6):
        left_list.append(tuple([" " + str(x+1), '', '', '']))
    
    for x in range(6, len(final_list)):
        temp = list(final_list[x])
        str_temp = str(x + 1)
        if x < 9:
            str_temp = " " + str(x+1)
        temp.insert(0, str_temp)
        right_list.append(temp)
    for x in range(max(6, len(final_list)), 12):
        str_temp = str(x + 1)
        if x < 9:
            str_temp = " " + str(x+1)
        right_list.append(tuple([str_temp, '', '', '']))
    return (left_list, right_list)



# code for the game logic

def store_player_list_nicer(new_list):
    global clean_player_list
    clean_player_list = list()
    for x in range(len(new_list)):
        item = new_list[x]
        clean_player_list.append(Player(x + 1, item[0], item[1], item[2] == 'true'))

# code for getting scoring values
def get_scoring_values():
    scoring_info = 'C:\\Users\\siniz\\AppData\\Local\\Programs\\Python\\Python38\\rhythm multiplayer 2\\static\\game_scoring\\Game_Scoring.xlsx'
    wookbook = openpyxl.load_workbook(scoring_info)
    scoring_dict = dict()

    # Define variable to read the active sheet:
    worksheet = wookbook.active

    # Iterate the loop to read the cell values
    # why can't i find 2p in col a?
    for i in range(2, worksheet.max_row+1):
        category_value = worksheet["{}{}".format('B', i)].value
       
        if (category_value != None and category_value == 'Regular'):
            scoring_dict[worksheet["{}{}".format('A', i)].value] = str(worksheet["{}{}".format('C', i)].value) +":" + str(worksheet["{}{}".format('D', i)].value)
        
        elif (category_value != None and category_value == 'Challenge'):
            scoring_dict["Challenge"] = str(worksheet["{}{}".format('E', i)].value) +":" + str(worksheet["{}{}".format('F', i)].value) + ":" + str(worksheet["{}{}".format('G', i)].value)

        elif(category_value != None and category_value == 'Endless'):
            scoring_dict[worksheet["{}{}".format('A', i)].value + ":Endless"] = str(worksheet["{}{}".format('H', i)].value) +":" + str(worksheet["{}{}".format('I', i)].value)
        
        elif (category_value != None and category_value == 'Vs'):
            scoring_dict[worksheet["{}{}".format('A', i)].value] = str(worksheet["{}{}".format('J', i)].value)
    
    return scoring_dict

# create all the info of the games
def get_filtered_games():
    BASE_DIR = 'C:\\Users\\siniz\\AppData\\Local\\Programs\\Python\\Python38\\rhythm multiplayer 2\\static\\rhythm_games'
    global global_file_data
    global global_game_category
    filtered_games = list()
    color_dict = dict()
    for category in global_game_category:
        color_dict[category] = global_game_category[category][1]

    index = 1
    file_list = convert_to_list(global_file_data)
    for file in file_list:
        print(file)
        full_file_name = BASE_DIR + '\\' + file
        wookbook = openpyxl.load_workbook(full_file_name)

        # Define variable to read the active sheet:
        worksheet = wookbook.active

        # Iterate the loop to read the cell values
        
        for row in range(2, worksheet.max_row+1):
            temp_list = list()
            for col in range(1, worksheet.max_column+1):
                cell = worksheet.cell(row=row, column=col)
                if (cell.value != None):
                    temp_list.append(cell.value)
            # all required attributes go up to 8 cols
            # anything after that is part of the "required games"
            print('nonfiltered' + str(temp_list))
            if (len(temp_list) >= 8):
                print(temp_list)
                required_list = list()
                if (len(temp_list) == 9):
                    required_list.append(temp_list[8])
                elif (len(temp_list) > 9):
                    required_list.extend(temp_list[8:])
                color_name = color_dict[temp_list[1]]


                filtered_games.append(Minigame(index, temp_list[0], temp_list[1], temp_list[5] == 'Yes', temp_list[3] == 'Yes', temp_list[2], temp_list[6], color_name, temp_list[7] == 'Yes', required_list))
                index+=1
    return filtered_games




#initialize game session
def initialize_game_session():
    global global_game_session
    global clean_player_list
    global global_settings_data
    global global_probability_data
    global_game_session = GameSession(clean_player_list, int(global_settings_data[0][2]), int(global_settings_data[0][0]), convert_to_list(global_probability_data), get_filtered_games(), get_scoring_values(), int(global_settings_data[0][1]))

# get a game with data
def get_game_turn_data(skipped_game: str):
    global global_game_session
    global global_current_game
    game, chosen_players_list = global_game_session.generate_data_for_turn(skipped_game)
    global_current_game = game
    next_players_list = list()
    next_players_list.append(global_game_session.get_next_player())
    game_list = convert_game_data_to_string_list(game)
    point_list = point_system_to_list(game, chosen_players_list)
    chosen_players_data_list = convert_player_data_to_list(chosen_players_list)
    next_players_data_list = convert_player_data_to_list(next_players_list)
    return (game_list, point_list, chosen_players_data_list, next_players_data_list)

def convert_game_data_to_string_list(game: Minigame) -> list:
    data_list = list()
    temp_list = list()
    temp_list.append(game.get_game_name()) # game name
    temp_list.append(game.get_category()) # game category
    temp_list.append(game.get_platform()) # game platform
    temp_list.append(game.get_color()) # color
    temp_list.append(game.get_title_card_url()) # image
    # this assumes that the category is not endless
    data_list.append(temp_list)
    return data_list

def point_system_to_list(game: Minigame, chosen_players_list: list) -> list:
    global global_game_session
    if (game.get_category() != 'Endless'):
        data_list = [float(item) for item in global_game_session.get_scoring()[game.get_category()].split(":")] # points
    else:
        data_list = global_game_session.get_scoring()[gaming.get_name() + ":Endless"].split(":")
    if len(chosen_players_list) == 2 and (chosen_players_list[0].get_handicap() or chosen_players_list[1].get_handicap()):
        data_list.append(True) # append handicapOn attribute
    elif len(chosen_players_list) == 1 and chosen_players_list[0].get_handicap():
        data_list.append(True)
    else:
        data_list.append(False)
    final_list = list()
    final_list.append(data_list)
    return final_list

def convert_player_data_to_list(chosen_players_list: list) -> list:
    data_list = list()
    if len(chosen_players_list) == 2:
        temp_list = list()
        temp_list.append(chosen_players_list[0].get_name() + "'s " + chosen_players_list[1].get_name() + "'s")
        for player in chosen_players_list:
            temp_list.append(player.get_pfp()) # pfp
            temp_list.append(player.get_handicap()) #handicap
        data_list.append(temp_list)
    else:
        player = chosen_players_list[0]
        temp_list = list()
        temp_list.append(player.get_name() + "'s") # name
        temp_list.append(player.get_pfp()) # pfp
        temp_list.append(player.get_handicap()) #handicap
        data_list.append(temp_list)
    return data_list





            


                

            




if __name__ == '__main__':
    rhm_site.run(debug=True)